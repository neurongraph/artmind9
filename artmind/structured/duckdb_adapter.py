"""DuckDB-over-parquet implementation of the ``Datasource`` connector.

xlsx/xlsm are read via ``openpyxl`` (bundled dep) rather than DuckDB's
``excel`` extension, which needs a network install at first use — that would
violate the no-network rule for hermetic tests (see plan Phase 1 risk (a)).
An excel sheet is staged to a temp CSV and loaded through the same
``read_csv_auto`` path as a native csv, so both formats infer types
identically.
"""

import csv
import tempfile
from pathlib import Path

import duckdb

import paths
from artmind.structured import view_name
from artmind.structured.connector import Column
from artmind.structured.profile import profile_column


def structured_db_path() -> Path:
    return paths.STRUCTURED_DIR / "artmind.duckdb"


def parquet_path_for(domain: str, table: str) -> Path:
    return paths.STRUCTURED_DIR / domain / f"{table}.parquet"


def _sql_quote_literal(path: Path) -> str:
    return "'" + str(path).replace("'", "''") + "'"


class DuckDBDatasource:
    #: Sentinel accepted by ``db_path`` to open an ephemeral, in-memory catalog
    #: instead of connecting to the shared persistent ``artmind.duckdb`` file.
    #: Callers that need query isolation (e.g. text2sql's domain-scoped
    #: execution) pass this explicitly rather than relying on a real path.
    IN_MEMORY = ":memory:"

    def __init__(self, db_path: Path | str | None = None):
        if db_path == self.IN_MEMORY:
            self.db_path = Path(self.IN_MEMORY)
            self.con = duckdb.connect(self.IN_MEMORY)
            return
        self.db_path = Path(db_path) if db_path else structured_db_path()
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.con = duckdb.connect(str(self.db_path))

    @classmethod
    def in_memory(cls) -> "DuckDBDatasource":
        """Open an ephemeral, in-memory catalog with no shared state.

        Used where a caller must guarantee that only explicitly-created views
        are reachable — e.g. text2sql's domain-scoped query execution, where
        the shared persistent catalog would still expose every other domain's
        already-ingested tables regardless of what ``ensure_views`` is called
        with.
        """
        return cls(db_path=cls.IN_MEMORY)

    def load_table(
        self,
        source: Path,
        table: str,
        domain: str,
        *,
        sheet: str | None = None,
        header_row: int = 0,
    ) -> int:
        source = Path(source)
        parquet = parquet_path_for(domain, table)
        parquet.parent.mkdir(parents=True, exist_ok=True)

        suffix = source.suffix.lower()
        cleanup: Path | None = None
        if suffix == ".csv":
            csv_path = source
        elif suffix in (".xlsx", ".xlsm"):
            csv_path = self._excel_sheet_to_csv(source, sheet=sheet)
            cleanup = csv_path
        else:
            raise ValueError(f"unsupported structured source: {source}")

        try:
            # DuckDB's COPY ... TO destination must be a literal, not a bind
            # parameter — the source can still be parameterized via read_csv_auto.
            self.con.execute(
                "COPY (SELECT * FROM read_csv_auto(?, header=true, skip="
                f"{int(header_row)})) TO {_sql_quote_literal(parquet)} (FORMAT PARQUET)",
                [str(csv_path)],
            )
        finally:
            if cleanup is not None:
                cleanup.unlink(missing_ok=True)

        self._create_view(table, parquet)
        return self.con.execute(
            "SELECT count(*) FROM read_parquet(?)", [str(parquet)]
        ).fetchone()[0]

    def stage_source(
        self,
        source: Path,
        *,
        sheet: str | None = None,
        header_row: int = 0,
    ) -> str:
        """Load a csv/xlsx source into a TEMP TABLE on this connection, without
        touching any target parquet file.

        Unlike ``load_table``, this never writes to ``parquet_path_for(...)`` —
        it's the staging primitive for a temporal refresh, where the incoming
        batch must be diffed against existing history (``scd2.apply_scd2_refresh``)
        rather than overwrite it outright. Reuses ``_excel_sheet_to_csv`` for
        xlsx so both formats go through the same ``read_csv_auto`` inference
        path as ``load_table``. Returns the temp table name, usable directly
        as a bare FROM-clause relation (e.g. ``apply_scd2_refresh``'s
        ``incoming_rel``).
        """
        source = Path(source)
        suffix = source.suffix.lower()
        cleanup: Path | None = None
        if suffix == ".csv":
            csv_path = source
        elif suffix in (".xlsx", ".xlsm"):
            csv_path = self._excel_sheet_to_csv(source, sheet=sheet)
            cleanup = csv_path
        else:
            raise ValueError(f"unsupported structured source: {source}")

        table_name = "_staged_source"
        try:
            self.con.execute(
                f"CREATE OR REPLACE TEMP TABLE {table_name} AS "
                "SELECT * FROM read_csv_auto(?, header=true, skip="
                f"{int(header_row)})",
                [str(csv_path)],
            )
        finally:
            if cleanup is not None:
                cleanup.unlink(missing_ok=True)
        return table_name

    def _excel_sheet_to_csv(self, source: Path, *, sheet: str | None) -> Path:
        # Imported lazily: openpyxl ships only in the optional `[ingest]` extra.
        # This path is reached outside Click (e.g. dashboard_routes.py), so it
        # raises RuntimeError rather than a click.ClickException.
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Reading xlsx/xlsm needs the optional 'ingest' extra. "
                "Install it with: uv tool install 'artmind9[ingest]'"
            ) from exc

        wb = load_workbook(source, read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet else wb.active
            fd, tmp_name = tempfile.mkstemp(suffix=".csv")
            tmp_path = Path(tmp_name)
            with open(fd, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            return tmp_path
        finally:
            wb.close()

    def _create_view(self, table: str, parquet: Path) -> None:
        # DDL statements can't be prepared/bound in DuckDB — the path must be a literal.
        self.con.execute(
            f'CREATE OR REPLACE VIEW "{table}" AS SELECT * FROM read_parquet({_sql_quote_literal(parquet)})'
        )

    def ensure_views(self, tables: list[dict]) -> None:
        """Recreate a VIEW per registered parquet so a fresh connection sees every table.

        Each table always gets its domain-qualified view (``view_name``,
        collision-free). It also gets the bare ``table_name`` as a convenience
        alias, but only when that name is unambiguous within this call — if two
        of the given tables share a ``table_name`` (same or different domain),
        neither gets the bare alias; any stale bare view left behind by an
        earlier, single-table ``load_table`` call is dropped too, so a query
        against the ambiguous bare name fails loudly instead of silently
        reading whichever domain's data was loaded last.
        """
        name_counts: dict[str, int] = {}
        for t in tables:
            name_counts[t["table_name"]] = name_counts.get(t["table_name"], 0) + 1

        for t in tables:
            parquet = Path(t["parquet_path"])
            self._create_view(view_name(t["domain"], t["table_name"]), parquet)
            if name_counts[t["table_name"]] == 1:
                self._create_view(t["table_name"], parquet)
            else:
                self.con.execute(f'DROP VIEW IF EXISTS "{t["table_name"]}"')

    def introspect_schema(self, table: str) -> list[Column]:
        rows = self.con.execute(f'DESCRIBE SELECT * FROM "{table}"').fetchall()
        return [Column(name=r[0], dtype=r[1]) for r in rows]

    def run_sql(self, sql: str) -> list[dict]:
        cursor = self.con.execute(sql)
        if cursor.description is None:
            return []
        columns = [d[0] for d in cursor.description]
        return [dict(zip(columns, row, strict=True)) for row in cursor.fetchall()]

    def profile_columns(self, table: str) -> dict:
        return {
            col.name: profile_column(self.con, table, col.name, col.dtype)
            for col in self.introspect_schema(table)
        }
