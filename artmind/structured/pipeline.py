"""Structured-file ingestion: csv/xlsx -> parquet + registry rows (replace mode).

Mirrors ``artmind/ingest.py``'s sha256 dedup shape, but the unit of dedup is
per registered table rather than per document.
"""

import dataclasses
import json
from datetime import date
from pathlib import Path

import click
from loguru import logger

from artmind.ingest import _compute_sha256
from artmind.structured import registry, sanitize_identifier
from artmind.structured.duckdb_adapter import DuckDBDatasource, parquet_path_for, structured_db_path
from artmind.structured.scd2 import SYSTEM_COLUMNS, apply_scd2_refresh

DATASOURCE_NAME = "default"


def _quote_ident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def _quote_literal(value: str) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _sheet_is_empty(ws) -> bool:
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            return False
    return True


def _enumerate_source_tables(
    source: Path, *, table: str | None, sheet: str | None, header_row: int
) -> list[dict]:
    """Return ``[{"table_name": ..., "sheet": sheet_or_None}]`` for the file."""
    if source.suffix.lower() == ".csv":
        return [{"table_name": sanitize_identifier(table or source.stem), "sheet": None}]

    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet_names = [
            name
            for name in wb.sheetnames
            if wb[name].sheet_state == "visible" and not _sheet_is_empty(wb[name])
        ]
    finally:
        wb.close()

    if sheet:
        if sheet not in sheet_names:
            raise click.ClickException(
                f"sheet '{sheet}' not found (or empty/hidden) in '{source.name}'"
            )
        sheet_names = [sheet]

    if not sheet_names:
        raise click.ClickException(f"no non-empty, visible sheets found in '{source.name}'")

    if len(sheet_names) == 1:
        return [{"table_name": sanitize_identifier(table or source.stem), "sheet": sheet_names[0]}]

    stem = sanitize_identifier(source.stem)
    return [
        {"table_name": f"{stem}__{sanitize_identifier(s)}", "sheet": s} for s in sheet_names
    ]


def _header_row_values(source: Path, sheet: str | None, header_row: int) -> list | None:
    if source.suffix.lower() == ".csv":
        import csv

        with open(source, newline="") as f:
            reader = csv.reader(f)
            for _ in range(header_row):
                next(reader, None)
            return next(reader, None)

    from openpyxl import load_workbook

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        rows = ws.iter_rows(values_only=True)
        for _ in range(header_row):
            next(rows, None)
        row = next(rows, None)
        return list(row) if row is not None else None
    finally:
        wb.close()


def _validate_header(source: Path, sheet: str | None, header_row: int) -> None:
    """Genuinely messy sheets (no clean header row) error out with guidance —
    v1 assumes reasonably tabular sheets (spec §6.1)."""
    header = _header_row_values(source, sheet, header_row)
    where = f"'{source.name}'" + (f" sheet '{sheet}'" if sheet else "")
    if not header or any(h is None or str(h).strip() == "" for h in header):
        raise click.ClickException(
            f"{where} has no clean header row (empty/missing column names)."
            " Clean the file first — see the `xlsx` skill."
        )


def _seed_temporal_history(
    ds: DuckDBDatasource,
    parquet_path: Path,
    business_key: str | None,
    effective_date_column: str | None,
) -> int:
    """First-time temporal seed: add ``_valid_from``/``_valid_to``/``_is_current``
    to a table that was just written as a raw (non-temporal) snapshot by
    ``ds.load_table``. Every row starts open (``_valid_to`` = NULL,
    ``_is_current`` = true); ``_valid_from`` follows the same effective-date
    rule as ``scd2.apply_scd2_refresh``'s refresh path -- the row's
    ``effective_date_column`` value if present and non-null, else today's
    date. Returns the (unchanged) row count.

    The existing history is materialized into a TEMP TABLE before the parquet
    is overwritten, mirroring ``scd2.apply_scd2_refresh``'s own read-then-copy
    discipline (never reading and overwriting the same parquet file in a
    single statement).
    """
    if not business_key:
        raise ValueError("temporal refresh_mode requires a business_key to seed history for")

    con = ds.con
    parquet_literal = _quote_literal(str(parquet_path))

    # Catch a typo'd --businessKey/--effectiveDateColumn now, at seed time,
    # rather than silently seeding an unrefreshable table and only surfacing
    # the mistake at the next `db refresh` (mirrors scd2.py's own
    # missing_bk check).
    cur_columns = {
        r[0] for r in con.execute(f"DESCRIBE SELECT * FROM read_parquet({parquet_literal})").fetchall()
    }
    key_columns = [c.strip() for c in business_key.split(",") if c.strip()]
    missing_bk = [c for c in key_columns if c not in cur_columns]
    if missing_bk:
        raise click.ClickException(f"--businessKey column(s) not found in source: {missing_bk}")
    if effective_date_column and effective_date_column not in cur_columns:
        raise click.ClickException(
            f"--effectiveDateColumn '{effective_date_column}' not found in source"
        )

    batch_date = date.today().isoformat()
    eff_expr = (
        f"COALESCE({_quote_ident(effective_date_column)}, DATE {_quote_literal(batch_date)})"
        if effective_date_column
        else f"DATE {_quote_literal(batch_date)}"
    )
    con.execute(
        "CREATE OR REPLACE TEMP TABLE _seed_temporal AS "
        f"SELECT *, {eff_expr} AS _valid_from, CAST(NULL AS DATE) AS _valid_to,"
        f" true AS _is_current FROM read_parquet({parquet_literal})"
    )
    con.execute(f"COPY _seed_temporal TO {parquet_literal} (FORMAT PARQUET)")
    return con.execute(f"SELECT count(*) FROM read_parquet({parquet_literal})").fetchone()[0]


def _register_columns_and_mappings(
    ds: DuckDBDatasource, table_id: int, table_name: str, domain: str, *, exclude_system_cols: bool
) -> None:
    """Introspect + profile ``table_name``'s current schema, persist it to the
    registry's ``columns`` table, and best-effort propose column-to-entity-
    class mappings. Shared by the initial-ingest (``_write_table``) and
    temporal-refresh (``_refresh_temporal_table``) paths so this bookkeeping
    can't drift between them.

    ``exclude_system_cols`` must be true whenever ``table_name`` carries the
    SCD-2 system columns (``_valid_from``/``_valid_to``/``_is_current``) --
    they're internal bookkeeping, not real data columns and shouldn't be
    profiled at all. Any *real* data column can still have a DATE/DECIMAL/etc
    ``distinct_sample`` -- ``default=str`` below handles those.
    """
    profiles = ds.profile_columns(table_name)
    columns = [
        {
            "name": c.name,
            "dtype": c.dtype,
            "profile_json": json.dumps(dataclasses.asdict(profiles[c.name]), default=str)
            if c.name in profiles
            else None,
        }
        for c in ds.introspect_schema(table_name)
        if not (exclude_system_cols and c.name in SYSTEM_COLUMNS)
    ]
    registry.replace_columns(table_id, columns)

    # Best-effort: propose column-to-entity-class mappings from the domain KG.
    # A down/unreachable graph must not fail the load (mirrors commit_to_graph's
    # hook guarding in artmind/ingest.py).
    try:
        from artmind.structured.mappings import propose_mappings

        propose_mappings(table_id, [domain])
    except Exception as e:
        logger.warning("structured pipeline: mapping proposal failed for {}: {}", table_name, e)


def _validate_temporal_incoming_columns(
    ds: DuckDBDatasource, parquet_path: Path, staged_rel: str, table_name: str
) -> None:
    """Guard against schema drift between a temporal table's existing history
    and the incoming batch, before ``scd2.apply_scd2_refresh`` ever runs.

    Without this check, a dropped/added non-key column surfaces as a raw
    ``duckdb.BinderException`` from deep inside ``apply_scd2_refresh``'s
    ``hash(ROW(...))`` construction -- this raises a clear ``ValueError``
    instead, matching ``refresh_table``'s own error-reporting convention (the
    CLI's ``db refresh`` catches ``ValueError`` and re-raises as a
    ``ClickException``).
    """
    parquet_literal = _quote_literal(str(parquet_path))
    cur_columns = {
        r[0] for r in ds.con.execute(f"DESCRIBE SELECT * FROM read_parquet({parquet_literal})").fetchall()
    }
    expected = cur_columns - set(SYSTEM_COLUMNS)
    incoming = {r[0] for r in ds.con.execute(f"DESCRIBE {staged_rel}").fetchall()}

    missing = sorted(expected - incoming)
    extra = sorted(incoming - expected)
    if missing or extra:
        detail = []
        if missing:
            detail.append(f"missing column(s) {missing}")
        if extra:
            detail.append(f"unexpected new column(s) {extra}")
        raise ValueError(
            f"table '{table_name}': incoming source's columns don't match the existing"
            f" temporal history ({'; '.join(detail)}) -- a temporal refresh requires the"
            " same column set as the seeded history; reconcile the columns first, or"
            " re-ingest as a fresh table if the schema is meant to change"
        )


def _write_table(
    ds: DuckDBDatasource,
    source: Path,
    domain: str,
    spec: dict,
    file_sha256: str,
    header_row: int,
    *,
    refresh_mode: str = "replace",
    business_key: str | None = None,
    effective_date_column: str | None = None,
) -> dict:
    _validate_header(source, spec["sheet"], header_row)
    row_count = ds.load_table(
        source, spec["table_name"], domain, sheet=spec["sheet"], header_row=header_row
    )
    parquet_path = parquet_path_for(domain, spec["table_name"])

    if refresh_mode == "temporal":
        row_count = _seed_temporal_history(ds, parquet_path, business_key, effective_date_column)

    table_id = registry.register_table(
        DATASOURCE_NAME,
        spec["table_name"],
        domain,
        source_file=str(source),
        sheet=spec["sheet"],
        parquet_path=str(parquet_path),
        row_count=row_count,
        sha256=file_sha256,
        refresh_mode=refresh_mode,
        business_key=business_key,
        effective_date_column=effective_date_column,
    )
    _register_columns_and_mappings(
        ds, table_id, spec["table_name"], domain, exclude_system_cols=(refresh_mode == "temporal")
    )
    table_row = registry.get_table(spec["table_name"], domain=domain)

    return {
        "table_name": spec["table_name"],
        "domain": domain,
        "row_count": row_count,
        "parquet_path": str(parquet_path),
        "version": table_row["version"],
    }


def _project_catalogue_best_effort(domain: str) -> None:
    """Re-project the Neo4j catalogue subgraph for ``domain`` after a table
    write. A down/unreachable graph must not fail the load (mirrors
    commit_to_graph's hook guarding in artmind/ingest.py, and the mapping-
    proposal hook above)."""
    try:
        from artmind.structured.catalogue import project_catalogue

        project_catalogue(domain)
    except Exception as e:
        logger.warning("structured pipeline: catalogue projection failed for domain '{}': {}", domain, e)


def ingest_structured_file(
    source: Path,
    domain: str,
    *,
    table: str | None = None,
    sheet: str | None = None,
    header_row: int = 0,
    force: bool = False,
    refresh_mode: str = "replace",
    business_key: str | None = None,
    effective_date_column: str | None = None,
) -> dict:
    if refresh_mode == "temporal" and not business_key:
        raise click.ClickException(
            "refresh_mode='temporal' requires --businessKey (business_key) —"
            " a temporal table can never be refreshed again without one"
        )

    source = Path(source)
    file_sha256 = _compute_sha256(source)

    registry.register_datasource(DATASOURCE_NAME, "duckdb", str(structured_db_path()))

    table_specs = _enumerate_source_tables(source, table=table, sheet=sheet, header_row=header_row)

    if not force:
        existing_rows = [
            registry.get_table(spec["table_name"], domain=domain) for spec in table_specs
        ]
        if existing_rows and all(
            row and row.get("sha256") == file_sha256 for row in existing_rows
        ):
            return {
                "status": "skipped",
                "domain": domain,
                "tables": [row["table_name"] for row in existing_rows],
            }

    ds = DuckDBDatasource()
    results = [
        _write_table(
            ds,
            source,
            domain,
            spec,
            file_sha256,
            header_row,
            refresh_mode=refresh_mode,
            business_key=business_key,
            effective_date_column=effective_date_column,
        )
        for spec in table_specs
    ]

    _project_catalogue_best_effort(domain)

    return {"status": "ok", "tables": results}


def _refresh_temporal_table(
    ds: DuckDBDatasource, existing: dict, source: Path, domain: str, file_sha256: str
) -> dict:
    """Diff ``source`` against a temporal table's existing SCD-2 history and
    rewrite the parquet with the merged full history (``scd2.apply_scd2_refresh``
    does the actual diff/rewrite). Unlike ``_write_table``, this never calls
    ``ds.load_table`` against the target parquet -- that would discard prior
    history wholesale. The incoming batch is staged into a TEMP TABLE instead
    (``DuckDBDatasource.stage_source``) and passed to ``apply_scd2_refresh`` as
    its ``incoming_rel``.
    """
    table_name = existing["table_name"]
    business_key = [c.strip() for c in (existing["business_key"] or "").split(",") if c.strip()]
    if not business_key:
        raise ValueError(
            f"table '{table_name}' has refresh_mode='temporal' but no business_key recorded"
        )

    parquet_path = Path(existing["parquet_path"])
    staged_rel = ds.stage_source(source, sheet=existing["sheet"], header_row=0)
    _validate_temporal_incoming_columns(ds, parquet_path, staged_rel, table_name)

    effective_date_column = existing["effective_date_column"]
    scd2_result = apply_scd2_refresh(
        ds.con,
        parquet_path,
        staged_rel,
        business_key=business_key,
        effective_date=date.today().isoformat(),
        effective_date_column=effective_date_column,
    )

    row_count = ds.con.execute(
        f"SELECT count(*) FROM read_parquet({_quote_literal(str(parquet_path))})"
    ).fetchone()[0]

    table_id = registry.register_table(
        DATASOURCE_NAME,
        table_name,
        domain,
        source_file=str(source),
        sheet=existing["sheet"],
        parquet_path=str(parquet_path),
        row_count=row_count,
        sha256=file_sha256,
        refresh_mode="temporal",
        business_key=existing["business_key"],
        effective_date_column=effective_date_column,
    )

    _register_columns_and_mappings(ds, table_id, table_name, domain, exclude_system_cols=True)
    table_row = registry.get_table(table_name, domain=domain)

    return {
        "table_name": table_name,
        "domain": domain,
        "row_count": row_count,
        "parquet_path": str(parquet_path),
        "version": table_row["version"],
        "scd2": scd2_result,
    }


def refresh_table(table_name: str, domain: str) -> dict:
    """Re-run the load for an already-registered table from its recorded
    ``source_file``. Bumps ``version``.

    ``replace`` mode overwrites the parquet with the current snapshot
    (``_write_table``, same as a fresh ingest). ``temporal`` mode diffs the
    incoming batch against the existing SCD-2 history and rewrites the parquet
    with the merged full history (``_refresh_temporal_table``), leaving prior
    versions of changed/removed rows intact.
    """
    existing = registry.get_table(table_name, domain=domain)
    if not existing:
        raise ValueError(f"table '{table_name}' is not registered for domain '{domain}'")
    if not existing.get("source_file"):
        raise ValueError(f"table '{table_name}' has no recorded source_file to refresh from")

    source = Path(existing["source_file"])
    file_sha256 = _compute_sha256(source)
    ds = DuckDBDatasource()

    if existing["refresh_mode"] == "temporal":
        result = _refresh_temporal_table(ds, existing, source, domain, file_sha256)
    else:
        spec = {"table_name": table_name, "sheet": existing["sheet"]}
        result = _write_table(ds, source, domain, spec, file_sha256, header_row=0)

    _project_catalogue_best_effort(domain)

    return {"status": "ok", **result}
